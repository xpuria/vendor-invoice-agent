"""Deterministic decision logic: approved-supplier matching + validation.

Supplier matching delegates the fuzzy *match* to a PydanticAI agent (legal-suffix and
parenthetical variants), but loads the list deterministically. Validation then applies the
brief's rules verbatim — no LLM, no regex — and decides post vs flag. The LLM never makes the
post/flag call; `decide` does.
"""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from pydantic_ai import Agent

from app.config import ROOT, models
from app.models import Decision, ExtractionResult, InvoiceData, ValidationResult


# approved-supplier matching
@dataclass(frozen=True)
class Supplier:
    supplier_id: str
    name: str
    status: str
    currency: str | None = None  # expected currency from the supplier master


@lru_cache
def load_suppliers(path: str | None = None) -> tuple[Supplier, ...]:
    xlsx = Path(path) if path else ROOT / "data" / "approved_suppliers.xlsx"
    ws = load_workbook(xlsx, read_only=True, data_only=True).active
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    h = next(i for i, r in enumerate(rows) if r and r[0] == "Supplier ID")
    header = [str(c).strip() if c else "" for c in rows[h]]
    id_i, name_i, status_i = (
        header.index("Supplier ID"),
        header.index("Vendor (Partner) Name"),
        header.index("Status"),
    )
    ccy_i = header.index("Currency") if "Currency" in header else None

    def cell(row, i):
        return row[i] if i is not None and i < len(row) else None

    out = []
    for r in rows[h + 1 :]:
        sid, name = cell(r, id_i), cell(r, name_i)
        if not sid or not name or not str(sid).startswith("SUP-"):
            continue
        ccy = cell(r, ccy_i)
        out.append(
            Supplier(str(sid), str(name).strip(), str(cell(r, status_i) or "").strip(),
                     str(ccy).strip() if ccy else None)
        )
    return tuple(out)


class SupplierMatch(BaseModel):
    """The LLM's verdict on whether a vendor is on the approved list."""

    supplier_id: str | None = Field(
        None, description="Supplier ID of the matched approved supplier, or null if none."
    )
    matched_name: str | None = None
    confidence: float = Field(0.0, ge=0, le=1, description="0..1 match confidence.")
    reasoning: str = ""


_MATCH_INSTRUCTIONS = (
    "You match a vendor name extracted from an invoice against a fixed Approved Supplier "
    "List. Account for legal-suffix variants (Ltd/Inc/SARL/Pty), parentheticals "
    "(e.g. 'Functional Software, Inc. (Sentry)'), and product-brand vs legal-entity naming "
    "(e.g. 'Sentry' == 'Functional Software'). Only return a supplier_id if you are "
    "confident it is the SAME legal entity; otherwise return null. Never invent an ID that "
    "is not in the provided list."
)


@lru_cache
def _agent(model: str) -> Agent:
    return Agent(model, output_type=SupplierMatch, instructions=_MATCH_INSTRUCTIONS)


@dataclass
class MatchResult:
    matched_name: str | None
    supplier_id: str | None
    score: float           # 0..100 for parity with the rest of the pipeline
    approved: bool
    reasoning: str
    currency: str | None = None  # expected currency from the supplier master


def match(vendor_name: str | None, path: str | None = None) -> MatchResult:
    if not vendor_name or not vendor_name.strip():
        return MatchResult(None, None, 0.0, False, "no vendor name extracted")
    suppliers = load_suppliers(path)
    listing = "\n".join(f"{s.supplier_id}\t{s.name}\t[{s.status}]" for s in suppliers)
    by_id = {s.supplier_id: s for s in suppliers}
    prompt = (
        f"Approved Supplier List (id, name, status):\n{listing}\n\n"
        f"Invoice vendor name: {vendor_name!r}\n"
        "Which approved supplier is this, if any?"
    )
    verdict = _agent(models()["primary"]).run_sync(prompt).output
    sup = by_id.get(verdict.supplier_id) if verdict.supplier_id else None
    approved = bool(sup and sup.status.lower() == "approved")
    return MatchResult(
        matched_name=sup.name if sup else None,
        supplier_id=sup.supplier_id if sup else None,
        score=verdict.confidence * 100,
        approved=approved,
        reasoning=verdict.reasoning,
        currency=sup.currency if sup else None,
    )


# validation — rules taken verbatim from the brief

# the seven required fields from the brief's "data to extract" table
MANDATORY = [
    "vendor_name", "invoice_number", "invoice_date", "total_amount",
    "line_items", "payment_terms", "due_date",
]


def check_complete(data: InvoiceData) -> tuple[bool, list[str]]:
    """An invoice is valid only when all mandatory fields are present (brief, literal)."""
    missing = []
    for f in MANDATORY:
        val = getattr(data, f)
        if f == "line_items":
            if not val:
                missing.append(f)
        elif val is None or (isinstance(val, str) and not val.strip()):
            missing.append(f)
    return (not missing, missing)


def decide(extraction: ExtractionResult, supplier: MatchResult, is_duplicate: bool,
           confidence_threshold: float = 0.7) -> ValidationResult:
    """Precedence: duplicate -> low-confidence -> incomplete -> unapproved -> discrepancy
    -> AUTO_POST."""
    data = extraction.data
    is_complete, missing = check_complete(data)
    reasons: list[str] = []
    base = dict(
        is_complete=is_complete, missing_fields=missing,
        vendor_approved=supplier.approved, matched_supplier=supplier.matched_name,
        match_score=supplier.score, is_duplicate=is_duplicate,
    )

    if is_duplicate:
        reasons.append(
            f"Duplicate of an already-seen bill ({data.vendor_name} / {data.invoice_number}).")
        return ValidationResult(decision=Decision.FLAG_DUPLICATE, reasons=reasons, **base)

    # "suspicious" in the brief — low confidence routes to a human
    conf = extraction.overall_confidence
    if conf is not None and conf < confidence_threshold:
        reasons.append(f"Extraction confidence {conf:.2f} below {confidence_threshold:.2f}.")
        return ValidationResult(decision=Decision.FLAG_LOW_CONFIDENCE, reasons=reasons, **base)

    if not is_complete:
        reasons.append("Missing mandatory field(s): " + ", ".join(missing) + ".")
        return ValidationResult(decision=Decision.FLAG_INCOMPLETE, reasons=reasons, **base)

    if not supplier.approved:
        reasons.append(
            f"Vendor '{data.vendor_name}' not on Approved Supplier List "
            f"(best match score {supplier.score:.0f}) — route to Supplier Evaluation.")
        return ValidationResult(decision=Decision.FLAG_NEW_VENDOR, reasons=reasons, **base)

    # cross-check currency vs the supplier master; a mismatch is ambiguity or an anomaly —
    # escalate to a human rather than auto-posting on a guess
    got_ccy = (data.currency or "").strip().upper()
    master_ccy = (supplier.currency or "").strip().upper()
    if got_ccy and master_ccy and got_ccy != master_ccy:
        reasons.append(
            f"Currency discrepancy: invoice shows {got_ccy}, supplier master expects "
            f"{master_ccy}. Confirm before posting.")
        return ValidationResult(decision=Decision.FLAG_DISCREPANCY, reasons=reasons, **base)

    reasons.append(
        f"Complete and vendor approved ({supplier.matched_name}, score {supplier.score:.0f}).")
    return ValidationResult(decision=Decision.AUTO_POST, reasons=reasons, **base)
